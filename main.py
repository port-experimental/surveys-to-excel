import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from pathlib import Path
import re

SURVEYS_DIR = "surveys"
EXCEL_PATH = "all_surveys.xlsx"

headers = ["Field", "Title", "Type", "Enum", "Description"]

excel_writer = pd.ExcelWriter(EXCEL_PATH, engine="openpyxl")

for survey_file in Path(SURVEYS_DIR).glob("*.json"):
    with open(survey_file, "r") as f:
        data = json.load(f)
    # Defensive: find the properties field
    try:
        properties = data["trigger"]["userInputs"]["properties"]
    except KeyError:
        continue
    rows = []
    for key, prop in properties.items():
        # Check for nested 'items' property with 'enum'
        if (
            "items" in prop
            and isinstance(prop["items"], dict)
            and "enum" in prop["items"]
        ):
            enum_list = prop["items"]["enum"]
        else:
            enum_list = prop.get("enum", []) if "enum" in prop else []
        enum_str = ",".join(str(e) for e in enum_list)
        row = {
            "Field": key,
            "Title": prop.get("title", ""),
            "Type": prop.get("type", ""),
            "Enum": enum_str,  # This will be used for dropdown
            "Description": prop.get("description", ""),
        }
        rows.append(row)
    df = pd.DataFrame(rows, columns=headers)
    # Use the survey title or filename as the sheet name
    sheet_name = data.get("title") or survey_file.stem
    # Replace invalid Excel sheet characters (\ / ? * [ ] :) with '-'
    sheet_name = re.sub(r"[\\/?*\[\]:]", "-", sheet_name)
    # Excel sheet names max 31 chars
    sheet_name = sheet_name[:31]
    df.to_excel(excel_writer, sheet_name=sheet_name, index=False)

excel_writer.close()

# Add data validation for enum fields in each sheet
wb = load_workbook(EXCEL_PATH)
for ws in wb.worksheets:
    enum_col_idx = headers.index("Enum") + 1
    enum_col_letter = get_column_letter(enum_col_idx)
    for idx, row in enumerate(
        ws.iter_rows(min_row=2, max_col=len(headers), values_only=True), start=2
    ):
        enum_val = row[headers.index("Enum")]
        if enum_val:
            dv = DataValidation(type="list", formula1=f'"{enum_val}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{enum_col_letter}{idx}")
            ws[f"{enum_col_letter}{idx}"].value = ""
wb.save(EXCEL_PATH)
print(f"Excel file '{EXCEL_PATH}' created with all survey properties as separate tabs.")
